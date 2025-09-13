# Passo a Passo:
"""

1. Fazer login no site (Selenium)
2. Pegar todos XPATH e criar várival
3. Abrir a planilha (Openpyxl)
4. Separar as informações e percorrer cada uma dela
5. Preencher cada campo com as informações da planilha
6. Emitir a nota
"""

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# Configurando para caso tenha notificação de manter download ou descartar
# Linhas de códigos para permitir
options = webdriver.ChromeOptions()
options.add_experimental_option("prefs", {
  "download.default_directory": r"C:\Users\Alan\Downloads",
  "download.prompt_for_download": False,
  "download.directory_upgrade": True,
  "safebrowsing.enabled": True
})

navegador = webdriver.Chrome(options=options)
navegador.get(r'C:\Users\Alan\Downloads\Arquivos-20220419T044739Z-001\Arquivos\login.html')

# Pegando informações dos campos:
login = navegador.find_element(By.XPATH, '/html/body/div/form/input[1]')
login.send_keys('corporativo@gmail.com')
sleep(0.5)

senha = navegador.find_element(By.XPATH, '/html/body/div/form/input[2]')
senha.send_keys('senha123')
sleep(0.5)

botao_login = navegador.find_element(By.XPATH, '/html/body/div/form/button')
botao_login.click()
sleep(0.5)

campo_nome = navegador.find_element(By.XPATH,'//*[@id="nome"]')
campo_endereço = navegador.find_element(By.XPATH, '//input[@name="endereco"]') # Pegando Xpath e informando o nome da clase e valor 
campo_bairro = navegador.find_element(By.XPATH, '//*[@id="formulario"]/input[3]')
campo_municipio = navegador.find_element(By.XPATH, '//*[@id="formulario"]/input[4]')
campo_cep = navegador.find_element(By.XPATH,'//*[@id="formulario"]/input[5]')
campo_uf = navegador.find_element(By.XPATH, '//*[@id="formulario"]/select')
campo_cpf_cnpj = navegador.find_element(By.XPATH, '//*[@id="formulario"]/input[6]')
campo_incrição_estadual = navegador.find_element(By.XPATH, '//*[@id="formulario"]/input[7]')
campo_descrição = navegador.find_element(By.XPATH, '//*[@id="formulario"]/input[8]')
campo_quantidade = navegador.find_element(By.XPATH, '//*[@id="formulario"]/input[9]')
campo_valor_unitario = navegador.find_element(By.XPATH, '//*[@id="formulario"]/input[10]')
campo_valor_total = navegador.find_element(By.XPATH, '//*[@id="formulario"]/input[11]')
botao_emitir_nota = navegador.find_element(By.XPATH, '//*[@id="formulario"]/button')

planilha_dados = load_workbook('NotasEmitir.xlsx') # Abrindo planilha
pagina_planilha = planilha_dados.active # Definindo a página atual

# Percorrendo a planilha, linha a linha (.iter_rows), começando pela segunda linha (min_row=2), pegando valores (values_only=True)
for linha in pagina_planilha.iter_rows(min_row=2, values_only=True): # Percorrendo a cada linha das colunas - min_row=2 (começar na linha 2) - values_only=True (Apenas valores)
    
    cliente = linha[0]
    cpf_cnpj = linha[1]
    cep = linha[2]
    endereço = linha[3]
    bairro = linha[4]
    municipio = linha[5]
    uf = linha[6]
    inscrição_estadual = linha[7]
    descrição = linha[8]
    quantidade = linha[9]
    valor_unitario = linha[10]
    valor_total = linha[11]
   
    campo_nome.send_keys(cliente)
    sleep(0.5)
    campo_endereço.send_keys(endereço)
    sleep(0.5)
    campo_bairro.send_keys(bairro)
    sleep(0.5)
    campo_municipio.send_keys(municipio)
    sleep(0.5)
    campo_cep.send_keys(cep)
    sleep(0.5)
    campo_uf.send_keys(uf)
    sleep(0.5)
    campo_cpf_cnpj.send_keys(cpf_cnpj)
    sleep(0.5)
    campo_incrição_estadual.send_keys(inscrição_estadual)
    sleep(0.5)
    campo_descrição.send_keys(descrição)
    sleep(0.5)
    campo_quantidade.send_keys(quantidade)
    sleep(0.5)
    campo_valor_unitario.send_keys(valor_unitario)
    sleep(0.5)
    campo_valor_total.send_keys(valor_total)
    sleep(0.5)
    botao_emitir_nota.click()
    sleep(1)
    
    '''  # Método de limpar todos os campos:
    # Encontra todos os campos de entrada do tipo texto, email, senha, etc.
    campos_clear = navegador.find_elements(By.XPATH, "//input[not(@type='hidden')]")
    for clear in campos_clear:   
        try:
            clear.clear() # Limpa todos os campos encontrados
        except:
            print(f'Não foi possivel limpar o campo: {clear}')
'''
    # 2. Metodo de limpar todos os campos
    lista_clear = navegador.find_elements(By.TAG_NAME, 'input')
    for clear in lista_clear:
        try:
            clear.clear()
        except:
            print(f'Não foi possivel limpar o campo: {clear}')


sleep(30)
